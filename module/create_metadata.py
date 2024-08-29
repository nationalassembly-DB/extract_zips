"""
process_folder에서 is_compressed_exists가 False인 경우 실행됩니다.
"""


import os
import pathlib
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from natsort import natsorted


from module.compress import error_files


def create_metadata(folder_path):
    """엑셀 파일과 리스트를 종합하여 엑셀파일을 생성합니다"""
    excel_path = os.path.join(folder_path, '파일리스트.xlsx')
    df = _dir_to_dic(folder_path)
    wb = _load_excel(excel_path)
    ws = wb.active
    _save_excel(df, ws, ws.max_row)
    wb.save(excel_path)


def _save_excel(df, ws, last_row):
    """엑셀파일에 df를 불러와 값을 입력합니다"""
    for index, row in df.iterrows():
        row_index = last_row + index + 1
        blank = str(row['위원회']).find(' ')
        under_bar = str(row['위원회']).find('_')
        if blank != -1 and under_bar != -1:
            cmt = str(row['위원회'])[blank+1:under_bar]
        elif blank != -1 and under_bar == -1:
            cmt = str(row['위원회'])[blank+1:]
        else:
            cmt = row['위원회']

        if row['피감기관'] == '파일리스트.xlsx':
            continue
        ws.cell(row=row_index, column=1, value=cmt)  # 위원회
        ws.cell(row=row_index, column=2, value=row['피감기관'])
        ws.cell(row=row_index, column=3, value=row['파일명'])
        ws.cell(row=row_index, column=4, value=row['경로'])
        ws.cell(row=row_index, column=5, value=row['확장자'])
        ws.cell(row=row_index, column=6, value=row['파일크기'])
        if row['파일크기'] == 0:
            ws.cell(row=row_index, column=7, value='0바이트')
        if row['실제경로'] in error_files:
            ws.cell(row=row_index, column=7, value=error_files[row['실제경로']])


def _load_excel(excel_file_path):
    """엑셀 파일을 불러옵니다. 파일이 존재하지 않은 경우 header를 추가하고 새 파일을 불러옵니다"""
    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active

        headers = ['위원회', '피감기관', '실제 파일명', '경로', '확장자', '파일크기', '특이사항']
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color
        wb.save(excel_file_path)
    wb = load_workbook(excel_file_path)

    return wb


def _dir_to_dic(folder_path):
    """폴더를 순회하여 파일명을 가져옵니다"""
    file_list = []

    for root, _, files in os.walk(folder_path):
        for file in natsorted(files):
            file_extension = pathlib.Path(file).suffix.lstrip('.').lower()
            file_path = os.path.join('\\\\?\\', root, file)
            relative_path = os.path.relpath(
                file_path, os.path.dirname(folder_path))

            file_list.append({
                '위원회': os.path.basename(folder_path),
                '피감기관': relative_path.split(os.sep)[1],
                '파일명': file,
                '경로': relative_path,
                '확장자': file_extension,
                '실제경로': file_path,
                '파일크기': os.path.getsize(file_path)
            })

    return pd.DataFrame(file_list)
